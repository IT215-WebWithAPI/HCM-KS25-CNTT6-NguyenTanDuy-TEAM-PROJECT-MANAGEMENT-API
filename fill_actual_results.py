from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

root = Path(__file__).parent
path = root / "TEAM_PROJECT_MANAGER_API_Test_Checklist.xlsm"
wb = load_workbook(path, keep_vba=True)
ws = wb["Checklist"]

results = {
    "TC_SYSTEM_01": ("Chưa thực thi request; cần MySQL hoạt động để xác nhận HTTP 200 và database=connected.", "Chưa test"),
    "TC_SYSTEM_02": ("Đã xác minh bằng import app.main trong venv: RuntimeError - Form data requires python-multipart to be installed; ứng dụng chưa khởi động được.", "Fail"),
    "TC_SYSTEM_03": ("Chưa thực thi; theo code create_all chạy lúc import nên DB lỗi có thể làm Uvicorn dừng với OperationalError.", "Chưa test"),
    "TC_AUTH_01": ("Chưa thực thi request; theo code khi đủ dependency và DB sẽ tạo User, nhưng cần xác nhận bằng HTTP 201 thực tế.", "Chưa test"),
    "TC_AUTH_02": ("Chưa thực thi request; code có nhánh HTTP 400 khi email đã tồn tại.", "Chưa test"),
    "TC_AUTH_03": ("Chưa thực thi request; request dùng Form và hiện bị chặn trước khi chạy do thiếu python-multipart trong venv.", "Blocked"),
    "TC_AUTH_04": ("Chưa thực thi request; theo code email sai định dạng sẽ đi vào RequestValidationError và dự kiến HTTP 422.", "Chưa test"),
    "TC_AUTH_05": ("Chưa thực thi request; theo code role ngoài admin/user sẽ bị từ chối bằng RequestValidationError.", "Chưa test"),
    "TC_AUTH_06": ("Chưa thực thi request; cần User tồn tại và DB hoạt động để xác nhận HTTP 200 cùng access_token/refresh_token.", "Chưa test"),
    "TC_AUTH_07": ("Chưa thực thi request; code trả HTTP 404 khi sai email hoặc password.", "Chưa test"),
    "TC_AUTH_08": ("Chưa thực thi request; cần refresh token hợp lệ và DB hoạt động.", "Chưa test"),
    "TC_AUTH_09": ("Chưa thực thi request; code có nhánh HTTP 401 cho JWT sai/hết hạn.", "Chưa test"),
    "TC_USER_01": ("Chưa thực thi request; cần access token hợp lệ và User trong DB.", "Chưa test"),
    "TC_USER_02": ("Chưa thực thi request; HTTPBearer dự kiến từ chối với HTTP 403/401 tùy cấu hình FastAPI, cần ghi nhận status thực tế.", "Chưa test"),
    "TC_USER_03": ("Chưa thực thi request; cần token admin và dữ liệu User.", "Chưa test"),
    "TC_USER_04": ("Chưa thực thi request; RoleChecker có nhánh HTTP 403 cho user không phải admin.", "Chưa test"),
    "TC_PROJECT_01": ("Chưa thực thi request; cần token hợp lệ và DB. Code tạo Project rồi tạo owner membership trong hai lần commit.", "Chưa test"),
    "TC_PROJECT_02": ("Chưa thực thi request; name_project rỗng được chuyển thành None rồi ProjectCreate có thể phát sinh ValidationError/HTTP 500 thay vì HTTP 422.", "Fail"),
    "TC_PROJECT_03": ("Chưa thực thi request; endpoint có dependency get_current_user nên dự kiến bị từ chối xác thực.", "Chưa test"),
    "TC_PROJECT_04": ("Chưa thực thi request; nếu User không có Project/member, service chủ động trả HTTP 404.", "Chưa test"),
    "TC_PROJECT_05": ("Chưa thực thi request; service có nhánh HTTP 404 khi Project không tồn tại hoặc không thuộc User.", "Chưa test"),
    "TC_PROJECT_06": ("Chưa thực thi request; schema ProjectUpdate hiện yêu cầu name/description nhưng router truyền None khi bỏ field, có thể ValidationError/HTTP 500.", "Fail"),
    "TC_PROJECT_07": ("Chưa thực thi request; service có nhánh HTTP 403 nếu current_user không phải owner.", "Chưa test"),
    "TC_MEMBER_01": ("Chưa thực thi request; lỗi member_input.role trước đây đã được bỏ trong code hiện tại, cần chạy lại để xác nhận HTTP 201.", "Chưa test"),
    "TC_MEMBER_02": ("Chưa thực thi request; service có nhánh HTTP 403 cho người không phải owner.", "Chưa test"),
    "TC_MEMBER_03": ("Chưa thực thi request; service có nhánh HTTP 400 khi user_id không tồn tại.", "Chưa test"),
    "TC_MEMBER_04": ("Chưa thực thi request; service có nhánh HTTP 400 và database có unique constraint cho thành viên trùng.", "Chưa test"),
    "TC_MEMBER_05": ("Chưa thực thi request; cần Project và dữ liệu member/user hợp lệ.", "Chưa test"),
    "TC_MEMBER_06": ("Chưa thực thi request; code serialize ORM member sau delete+commit, có nguy cơ lỗi object đã bị xóa/response không hợp lệ.", "Fail"),
    "TC_TASK_01": ("Chưa thực thi request; code hiện cho phép due_date=None và service đã guard khi tạo, nhưng cần DB/token/member hợp lệ để xác nhận HTTP 201.", "Chưa test"),
    "TC_TASK_02": ("Chưa thực thi request; service có nhánh HTTP 400 khi priority ngoài low/medium/high.", "Chưa test"),
    "TC_TASK_03": ("Chưa thực thi request; service có nhánh HTTP 400 khi due_date trong quá khứ.", "Chưa test"),
    "TC_TASK_04": ("Chưa thực thi request; service có nhánh HTTP 403 khi current_user không thuộc Project.", "Chưa test"),
    "TC_TASK_05": ("Chưa thực thi request; cần Project/member hợp lệ để xác nhận HTTP 200.", "Chưa test"),
    "TC_TASK_06": ("Chưa thực thi request; service có nhánh HTTP 404 nếu task_id không tồn tại và HTTP 403 nếu không có quyền.", "Chưa test"),
    "TC_TASK_07": ("Chưa thực thi request; service có nhánh HTTP 404 khi Task không tồn tại.", "Chưa test"),
    "TC_TASK_08": ("Đã xác minh schema runtime: PATCH chỉ gửi title tạo ValidationError vì TaskUpdate vẫn yêu cầu description, assignee_id, status, priority.", "Fail"),
    "TC_TASK_09": ("Chưa thực thi request; service có nhánh HTTP 400 cho status/priority không hợp lệ.", "Chưa test"),
    "TC_TASK_10": ("Chưa thực thi request; code update_task gọi .timestamp() nếu key due_date tồn tại, nên due_date=None có thể gây AttributeError và HTTP 500.", "Fail"),
    "TC_TASK_11": ("Chưa thực thi request; code trả ORM Task sau delete+commit rồi router serialize, có nguy cơ lỗi response.", "Fail"),
    "TC_TASK_12": ("Chưa thực thi request; service có nhánh HTTP 403 nếu member không phải owner.", "Chưa test"),
    "TC_TASK_13": ("Chưa thực thi request; có hai handler cùng GET /api/tasks, cần kiểm tra OpenAPI để biết handler thực tế.", "Fail"),
    "TC_TASK_14": ("Chưa thực thi request; service có nhánh HTTP 400 cho status/priority filter sai.", "Chưa test"),
    "TC_TASK_15": ("Chưa thực thi request; có hai handler cùng GET /api/tasks nên hành vi phân trang chưa xác định chắc chắn.", "Fail"),
    "TC_TASK_16": ("Chưa thực thi request; nếu handler phân trang được chọn, service có nhánh HTTP 400 cho sort_by/sort_order sai.", "Chưa test"),
    "TC_SECURITY_01": ("Chưa thực thi request; dependency JWT có nhánh HTTP 401 cho token sai, cần xác nhận response không lộ traceback.", "Chưa test"),
    "TC_SECURITY_02": ("Chưa thực thi request; hiện chưa thể mở OpenAPI vì import app.main bị chặn bởi python-multipart.", "Blocked"),
}

for row in range(2, ws.max_row + 1):
    test_id = ws.cell(row, 3).value
    if test_id in results:
        actual, status = results[test_id]
        ws.cell(row, 7).value = actual
        ws.cell(row, 8).value = status

# Make the distinction visible without changing the user's test data.
ws.conditional_formatting.add(f"H2:H{ws.max_row}", FormulaRule(formula=['H2="Pass"'], fill=PatternFill("solid", fgColor="C6EFCE")))
ws.conditional_formatting.add(f"H2:H{ws.max_row}", FormulaRule(formula=['H2="Fail"'], fill=PatternFill("solid", fgColor="FFC7CE")))
ws.conditional_formatting.add(f"H2:H{ws.max_row}", FormulaRule(formula=['H2="Blocked"'], fill=PatternFill("solid", fgColor="FFEB9C")))
wb.save(path)

# Keep the file recognized as XLSM by Excel.
tmp = path.with_suffix(".tmp")
with ZipFile(path, "r") as source, ZipFile(tmp, "w", ZIP_DEFLATED) as target:
    for item in source.infolist():
        content = source.read(item.filename)
        if item.filename == "[Content_Types].xml":
            content = content.replace(
                b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
                b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
            )
        target.writestr(item, content)
path.unlink()
tmp.rename(path)

check = load_workbook(path, read_only=True)
print("updated_rows=", sum(1 for row in range(2, check["Checklist"].max_row + 1) if check["Checklist"].cell(row, 7).value))
print("sheets=", check.sheetnames)
